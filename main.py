import openpyxl
from openpyxl.utils import get_column_letter
from itertools import combinations, product
from copy import copy
import re

def main():
    # Load the source workbook and select the active sheet
    source_wb = openpyxl.load_workbook('source.xlsx', data_only=True)
    source_ws = source_wb.active

    # Create a new workbook for the transformed data
    transformed_wb = openpyxl.Workbook()
    transformed_ws = transformed_wb.active

    # Copy the source data to the transformed sheet, preserving formats and styles
    for row_idx, row in enumerate(source_ws.iter_rows(), 1):
        for col_idx, cell in enumerate(row, 1):
            new_cell = transformed_ws.cell(row=row_idx, column=col_idx, value=cell.value)
            # Preserve date formatting for date headers (row 1)
            if row_idx == 1 and isinstance(cell.value, (float, int)):
                new_cell.number_format = 'YYYY-MM-DD'
            else:
                new_cell.number_format = cell.number_format
            # Copy styles
            new_cell.font = copy(cell.font)
            new_cell.fill = copy(cell.fill)
            new_cell.alignment = copy(cell.alignment)
            new_cell.border = copy(cell.border)
            new_cell.protection = copy(cell.protection)

    # Identify data starting row and column (after headers and labels)
    data_start_row = 3  # Adjust if your data starts on a different row
    data_start_col = 2  # Adjust if your data starts on a different column

    # Get the max row and column numbers
    max_row = transformed_ws.max_row
    max_col = transformed_ws.max_column

    # Dictionary to store evaluated cell values
    cell_values = {}

    # First pass: Store all cell values, treating empty cells as zero
    for col_idx in range(data_start_col, max_col + 1):
        for row_idx in range(data_start_row, max_row + 1):
            cell = transformed_ws.cell(row=row_idx, column=col_idx)
            cell_ref = (row_idx, col_idx)
            value = cell.value
            if value is None or (isinstance(value, str) and value.strip() == ''):
                value = 0.0
            else:
                value = parse_number(value)
            if value is not None:
                cell_values[cell_ref] = value
            else:
                cell_values[cell_ref] = None  # Non-numeric cell

    # Second pass: Process each bolded cell to find and apply formulas
    for col_idx in range(data_start_col, max_col + 1):
        for row_idx in range(data_start_row, max_row + 1):
            cell = transformed_ws.cell(row=row_idx, column=col_idx)

            # Check if the cell is bolded (should contain a formula)
            if cell.font and cell.font.bold and cell_values.get((row_idx, col_idx)) is not None:
                target_value = cell_values[(row_idx, col_idx)]

                # Find a formula by searching for arithmetic relationships
                formula = find_formula(transformed_ws, cell, data_start_row, cell_values)
                if formula:
                    cell.value = formula
                    # Update cell_values with the formula result
                    result = evaluate_expression(formula[1:], cell_values)  # Exclude '=' sign
                    if result is not None:
                        cell_values[(row_idx, col_idx)] = result

    # Save the transformed workbook
    transformed_wb.save('transformed.xlsx')
    print("Transformation complete. Saved as 'transformed.xlsx'.")

def parse_number(value):
    """
    Parse a number from a string that may contain commas and parentheses.
    """
    if isinstance(value, (int, float)):
        return float(value)
    elif isinstance(value, str):
        value = value.strip()
        # Remove commas
        value = value.replace(',', '')
        # Handle negative numbers in parentheses
        if value.startswith('(') and value.endswith(')'):
            value = '-' + value[1:-1]
        # Remove any additional formatting
        value = value.replace('$', '')  # Remove dollar signs if present
        try:
            return float(value)
        except ValueError:
            return None
    else:
        return None

def find_formula(ws, target_cell, data_start_row, cell_values, max_depth=10, tolerance=0.01):
    """
    Find a formula for the target cell by checking arithmetic relationships
    with other cells in the same column, using cells from top to bottom.
    """
    target_row = target_cell.row
    target_col = target_cell.column
    target_value = cell_values[(target_row, target_col)]

    # Collect previous cells (including formula cells) in the same column, from top to target_row
    candidate_cells = []
    for row in range(data_start_row, target_row):
        cell_ref = (row, target_col)
        value = cell_values.get(cell_ref)
        if value is not None:
            candidate_cells.append(cell_ref)

    # Limit the candidate cells to the last max_depth cells
    candidate_cells = candidate_cells[-max_depth:]

    n = len(candidate_cells)
    if n == 0:
        return None

    # Try combinations of candidate cells of size from 2 up to n
    for r in range(2, min(n + 1, max_depth + 1)):
        for combo in combinations(candidate_cells, r):
            # Generate all possible operator sequences for r - 1 operators
            ops_list = list(product(['+', '-'], repeat=r - 1))

            for ops in ops_list:
                # Build expression
                expr_elements = []
                for idx, cell_ref in enumerate(combo):
                    col_letter = get_column_letter(cell_ref[1])
                    cell_addr = f"{col_letter}{cell_ref[0]}"
                    if idx == 0:
                        # First operand without operator
                        expr_elements.append(cell_addr)
                    else:
                        expr_elements.append(ops[idx - 1] + cell_addr)
                expr = ''.join(expr_elements)
                # Evaluate the expression
                result = evaluate_expression(expr, cell_values)
                if result is not None and abs(result - target_value) <= tolerance:
                    # Construct the formula
                    formula = f"={expr}"
                    return formula

    return None

def evaluate_expression(expr, cell_values):
    """
    Evaluate the expression using the values from cell_values.
    """
    # Replace cell references with their values
    def replace_cell(match):
        cell_ref = match.group(0)
        col_letter, row_str = re.match(r'([A-Z]+)(\d+)', cell_ref).groups()
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
        row_idx = int(row_str)
        value = cell_values.get((row_idx, col_idx))
        if value is None:
            value = 0.0  # Treat empty or non-numeric cells as zero
        return str(value)

    expr_with_values = re.sub(r'[A-Z]+[0-9]+', replace_cell, expr)

    try:
        result = eval(expr_with_values)
        return result
    except Exception:
        return None

if __name__ == "__main__":
    main()
